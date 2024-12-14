import streamlit as st
import statistics
import base64
import io
import os
import matplotlib.pyplot as plt
import numpy as np
from tabulate import tabulate
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage
import io
import pandas as pd
import base64
from import_export import generate_settings_xlsx
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage
from PIL import Image as PILImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

def generate_analysis(chat_results, analyze_rating=True, analyze_length=True):
    total_cost = sum(chat_res["total_cost"] for chat_res in chat_results.values())

    # Prepare the analysis data table
    headers = ["Metric"] + [f"Chat {chat_index}" for chat_index in sorted(chat_results.keys())]
    analysis_data = [headers]

    if analyze_length:
        lengths_data = []
        for chat_index in sorted(chat_results.keys()):
            # Filter out None values from lengths
            lengths = [l for l in chat_results[chat_index]["lengths"] if l is not None]
            lengths_data.append(lengths)

        metrics = ["Average Length", "Median Length", "Std Dev Length", "Min Length", "Max Length"]
        stats_funcs = [statistics.mean, statistics.median, statistics.stdev, min, max]
        for metric, func in zip(metrics, stats_funcs):
            row = [metric]
            for lengths in lengths_data:
                if lengths:
                    if func == statistics.stdev and len(lengths) < 2:
                        value = "N/A"
                    else:
                        value = f"{func(lengths):.1f}"
                else:
                    value = "N/A"
                row.append(value)
            analysis_data.append(row)

    row = ["Iterations per chat"]
    for chat_index in sorted(chat_results.keys()):
        iterations = len(chat_results[chat_index]["responses"])
        row.append(str(iterations))
    analysis_data.append(row)

    if analyze_rating:
        ratings_data = []
        empty_rating_chats = []  # To track chats with no valid ratings
        for chat_index in sorted(chat_results.keys()):
            # Filter out None values from ratings
            ratings = [r for r in chat_results[chat_index]["ratings"] if r is not None]
            if not ratings:
                empty_rating_chats.append(f"Chat {chat_index}")
            ratings_data.append(ratings)

        # Print error if any chat has no valid ratings
        if empty_rating_chats:
            error_message = f"Error: No valid ratings found for {', '.join(empty_rating_chats)}."
            print(error_message)  # Prints to console
            st.error(error_message)  # Displays error in Streamlit app

        metrics = ["Average Rating", "95% CI of Rating", "SEM of Rating", "Std Dev of Rating", "Minimum Rating", "Maximum Rating"]
        stats_funcs = [
            statistics.mean,
            lambda x: f"±{1.96 * statistics.stdev(x) / np.sqrt(len(x)):.2f}",  # 95% CI
            lambda x: statistics.stdev(x) / np.sqrt(len(x)),  # SEM
            statistics.stdev,
            min,
            max
        ]
        for metric, func in zip(metrics, stats_funcs):
            row = [metric]
            for ratings in ratings_data:
                if ratings:
                    if func == statistics.stdev and len(ratings) < 2:
                        value = "N/A"
                    else:
                        try:
                            result = func(ratings)
                            # Check if result is already a string (like for CI function)
                            if isinstance(result, str):
                                value = result
                            else:
                                value = f"{result:.2f}"
                        except Exception:
                            value = "N/A"
                else:
                    value = "N/A"
                row.append(value)
            analysis_data.append(row)

    # Add total cost per chat
    row = ["Cost"]
    for chat_index in sorted(chat_results.keys()):
        cost = chat_results[chat_index]["total_cost"]
        row.append(f"${cost:.4f}")
    analysis_data.append(row)

    # Generate plots
    plot_base64 = ""
    plt.style.use('default')

    plot_cols = 0
    if analyze_length:
        plot_cols += 2
    if analyze_rating:
        plot_cols += 2

    if plot_cols == 0:
        # No plots to generate
        plot_base64 = ""
    else:
        fig, axs = plt.subplots(1, plot_cols, figsize=(5 * plot_cols, 6))
        if plot_cols == 1:
            axs = [axs]
        else:
            axs = axs.flatten()

        plot_idx = 0

        if analyze_length:
            # Length distribution histogram
            for chat_index in sorted(chat_results.keys()):
                lengths = chat_results[chat_index]["lengths"]
                # Filter out None values
                lengths = [l for l in lengths if l is not None]
                if lengths:
                    hist, bins = np.histogram(lengths, bins=10, density=True)
                    bin_centers = (bins[:-1] + bins[1:]) / 2
                    axs[plot_idx].fill_between(bin_centers, hist, alpha=0.5, label=f"Chat {chat_index}")
            axs[plot_idx].set_title('Response Length Distribution')
            axs[plot_idx].set_xlabel('Response Length (characters)')
            axs[plot_idx].set_ylabel('Density')
            axs[plot_idx].legend()
            plot_idx += 1

            # Length boxplot
            lengths_data = [ [l for l in chat_results[chat_index]["lengths"] if l is not None] for chat_index in sorted(chat_results.keys())]
            labels = [f"Chat {chat_index}" for chat_index in sorted(chat_results.keys())]
            axs[plot_idx].boxplot(lengths_data, labels=labels)
            axs[plot_idx].set_title('Response Length Box Plot')
            axs[plot_idx].set_ylabel('Response Length (characters)')
            plot_idx += 1

        if analyze_rating:
            # Rating distribution histogram
            max_rating = max(
                max(chat_results[chat_index]["ratings"]) for chat_index in chat_results if chat_results[chat_index]["ratings"] and any(r is not None for r in chat_results[chat_index]["ratings"])
            ) if any(chat_results[chat_index]["ratings"] and any(r is not None for r in chat_results[chat_index]["ratings"]) for chat_index in chat_results) else 5  # Default max rating
            bins = np.arange(0, int(max_rating) + 2)  # +2 to include the max value

            for idx, chat_index in enumerate(sorted(chat_results.keys())):
                ratings = [r for r in chat_results[chat_index]["ratings"] if r is not None]
                if ratings:
                    hist, _ = np.histogram(ratings, bins=bins)
                    # Normalize to get proportions
                    hist = hist / len(ratings)
                    x = np.arange(len(bins)-1)
                    bar_width = 0.8 / len(chat_results)
                    axs[plot_idx].bar(x + (idx * bar_width), hist, bar_width,
                                      label=f"Chat {chat_index}", alpha=0.7)
            if analyze_rating and any(chat_results[chat_index]["ratings"] and any(r is not None for r in chat_results[chat_index]["ratings"]) for chat_index in chat_results):
                axs[plot_idx].set_title('Rating Distribution')
                axs[plot_idx].set_xlabel('Rating')
                axs[plot_idx].set_ylabel('Proportion')
                axs[plot_idx].set_xticks(x)
                axs[plot_idx].set_xticklabels([f'{i:.1f}' for i in bins[:-1]])
                axs[plot_idx].legend()
            plot_idx += 1

            # Rating boxplot
            ratings_data_filtered = [ [r for r in chat_results[chat_index]["ratings"] if r is not None] for chat_index in sorted(chat_results.keys())]
            labels = [f"Chat {chat_index}" for chat_index in sorted(chat_results.keys())]
            axs[plot_idx].boxplot(ratings_data_filtered, labels=labels)
            axs[plot_idx].set_title('Rating Box Plot')
            axs[plot_idx].set_ylabel('Rating')
            plot_idx += 1

        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        plot_base64 = base64.b64encode(buf.getvalue()).decode('utf-8')
        plt.close()

    # Return analysis data, plot, and total cost
    return analysis_data, plot_base64, total_cost

def create_html_report(
    analysis_data,
    plot_base64,
    total_cost,
    chat_data,
    chat_results,
    model_response=None,
    model_rating=None,
    temperature_response=None,
    temperature_rating=None,
    evaluation_rubrics=None,
    analyze_rating=True,
    show_transcripts=False,
):
    # Construct combined HTML for each chat's system message, prompts, and evaluation rubric
    def extract_user_messages(messages):
        conversation = []
        for msg in messages:
            if msg['role'] == 'user':
                conversation.append(f"&#x1F9D1; {msg['content']}")
            elif msg['role'] == 'assistant':
                # Show [AI Responds] for blank assistant messages
                content = msg['content'].strip()
                conversation.append(f"&#x1F916; {content if content else '[AI Responds]'}")
        conversation_html = ('<br>').join(conversation)
        return conversation_html

    combined_html = ""

    for idx, chat_info in enumerate(chat_data, start=1):
        rubric_html = ""
        if analyze_rating and evaluation_rubrics and idx in evaluation_rubrics:
            rubric = evaluation_rubrics[idx]
            rubric_html = f"""
                <div class='response-box'>
                    {rubric}
                </div>
            """
        
        combined_html += f"""
        <div class="config-section" style="margin-bottom: 0;">
            <p><strong>Chat {idx}</strong></p>
            <div class="config-columns">
                <div class="config-column">
                    <div class="prompt-label">System Message:</div>
                    <div class="response-box" style="margin-bottom: 0;">{chat_info['system_message'] if chat_info['system_message'] else 'None'}</div>
                </div>
                <div class="config-column">
                    <div class="prompt-label" style="margin-top: 0;">Prompts:</div>
                    <div class="response-box">{extract_user_messages([msg for msg in chat_info['messages'] if msg['role'] == 'user' or msg['role'] == 'assistant'])}</div>
                </div>
            </div>
                <div class="prompt-label">Evaluation Rubric:</div>
                {rubric_html}
            </div>
        """

    # CSS styles with escaped curly braces
    css_styles = """
        body { font-family: Arial, sans-serif; margin: 40px; }
        table { border-collapse: collapse; width: 100%; margin: 20px 0; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }
        .response-box {
            background-color: #f8f9fa;
            border: 1px solid #ddd;
            padding: 15px;
            margin: 10px 0;
            border-radius: 4px;
        }
        .info-section {
            background-color: #f8f9fa;
            border: 1px solid #ddd;
            padding: 20px;
            margin: 20px 0;
            border-radius: 4px;
        }
        .config-columns {
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
        }
        .config-column {
            flex: 1;
            min-width: 250px;
        }
        .prompt-container {
            margin: 20px 0;
        }
        .prompt-label {
            font-weight: bold;
            color: #555;
            margin-bottom: 5px;
            display: block;
        }
        .config-section {
            margin-bottom: 40px;
        }
        h1, h2, h3, h4, h5 {
            color: #333;
        }
    """

    # Build the HTML content using f-strings
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Test Results</title>
        <style>
            {css_styles}
        </style>
    </head>
    <body>
        <h1>Research Results</h1>

        <div class="info-section">
            <h2>Configuration</h2>
            <p><strong>Number of Iterations:</strong> {len(next(iter(chat_results.values()))['responses']) if chat_results else 'N/A'}</p>
            <p><strong>Response Generation Model:</strong> {model_response if model_response else 'N/A'}</p>
            <p><strong>Response Temperature:</strong> {temperature_response if temperature_response else 'N/A'}</p>
            {f"<p><strong>Rating Model:</strong> {model_rating}</p>" if analyze_rating and model_rating else ""}
            {f"<p><strong>Rating Temperature:</strong> {temperature_rating}</p>" if analyze_rating and temperature_rating else ""}
            {f"<div class='response-box'><strong>Evaluation Rubric:</strong><br>{evaluation_rubric}</div>" if False else ""}  <!-- Removed global rubric display -->
            {combined_html}
        </div>

        <h2>Cost Analysis</h2>
        <p>Total API cost: ${total_cost:.4f}</p>

        <h2>Detailed Analysis</h2>
        {tabulate(analysis_data, headers='firstrow', tablefmt='html')}

        {f"<h2>Visualizations</h2><img src='data:image/png;base64,{plot_base64}' alt='Research Plots' style='max-width: 100%;'>" if plot_base64 else ""}
    """

    if show_transcripts and chat_results:
        transcripts_html = "<h2>Transcripts</h2>"

        # Sort chat_results by chat index to ensure Chat 1, Chat 2, etc.
        for chat_index in sorted(chat_results.keys()):
            results = chat_results[chat_index]
            transcripts_html += f"<h3>Chat {chat_index}</h3>"

            # Sort iterations to ensure Iteration 1, Iteration 2, etc.
            for iteration_idx, messages in enumerate(results["messages_per_iteration"], start=1):
                transcripts_html += f"<h4>Iteration {iteration_idx}</h4>"
                conversation_html = "<div class='response-box'>"
                for msg in messages:
                    role_symbol = '👤' if msg['role'] == 'user' else '🤖'
                    content = msg['content'].strip() if msg['content'].strip() else '[AI Responds]'
                    conversation_html += f"{role_symbol} {content}<br>"
                conversation_html += "</div>"
                transcripts_html += conversation_html

                if analyze_rating:
                    rating = results["ratings"][iteration_idx - 1] if results["ratings"] else "N/A"
                    rating_text = results["rating_texts"][iteration_idx - 1] if results["rating_texts"] else "N/A"
                    transcripts_html += f"""
                    <div class="response-box">&#x1F916; Rating: {rating} (raw response: "{rating_text}")</div>
                    """

        html_content += transcripts_html

    # Close the HTML tags
    html_content += """
    </body>
    </html>
    """

    return html_content

def generate_experiment_xlsx(
    settings_dict,
    chat_data,
    analysis_data,
    chat_results,
    plot_base64
):
    # Use the existing generate_settings_xlsx to create the initial BytesIO object
    initial_output = generate_settings_xlsx(settings_dict, chat_data, validate_chat=True)
    initial_output.seek(0)

    # Load the workbook from the BytesIO object using openpyxl
    workbook = load_workbook(initial_output)

    # Add Analysis Data sheet
    df_analysis = pd.DataFrame(analysis_data[1:], columns=analysis_data[0])
    sheet_analysis = workbook.create_sheet('Analysis Data')

    # Convert string values to numbers and apply formatting
    for r_idx, row in enumerate(dataframe_to_rows(df_analysis, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet_analysis.cell(row=r_idx, column=c_idx, value=value)
                
            # Skip header row and first column (metric names)
            if r_idx > 1 and c_idx > 1:
                if isinstance(value, str):
                    if value == 'N/A':
                        cell.value = 'N/A'
                    elif value.startswith('$'):
                        # Convert currency string to number
                        cell.value = float(value.replace('$', ''))
                        cell.number_format = '"$"#,##0.0000'
                    else:
                        # Convert other numeric strings to numbers
                        try:
                            cell.value = float(value)
                            # Use 2 decimal places for ratings, 1 for lengths
                            if 'Rating' in df_analysis.iloc[r_idx-1, 0]:
                                cell.number_format = '0.00'
                            else:
                                cell.number_format = '0.0'
                        except ValueError:
                            # Keep as string if conversion fails
                            pass

    # Adjust column widths
    for column_cells in sheet_analysis.columns:
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        sheet_analysis.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Create a consolidated Transcripts sheet
    sheet_transcripts = workbook.create_sheet('Transcripts')

    # Prepare data for transcripts
    columns_data = []
    max_messages = 0  # To track the maximum number of messages

    # Collect data per chat and iteration
    for chat_index in sorted(chat_results.keys()):
        chat_result = chat_results[chat_index]
        messages_per_iteration = chat_result['messages_per_iteration']
        ratings = chat_result.get('ratings', [])
        rating_texts = chat_result.get('rating_texts', [])

        for iteration_index, messages in enumerate(messages_per_iteration, 1):
            column = {
                'chat': f"Chat {chat_index}",
                'iteration': f"Iteration {iteration_index}",
                'messages': [],
                'rating_pos': None
            }

            # Collect messages
            for msg in messages:
                content = f"{msg['content']}"
                column['messages'].append(content)

            # Append rating at the end
            if ratings and iteration_index <= len(ratings):
                rating = ratings[iteration_index - 1]
                rating_text = rating_texts[iteration_index - 1] if rating_texts else ''
                rating_content = f"Rating: {rating} [verbatim rating: '{rating_text}']"
                column['rating_pos'] = len(column['messages']) + 2  # +2 to account for header rows
                column['messages'].append(rating_content)

            # Update max_messages
            if len(column['messages']) > max_messages:
                max_messages = len(column['messages'])

            columns_data.append(column)

    # Write headers
    sheet_transcripts.cell(row=1, column=1, value='Chat #')
    sheet_transcripts.cell(row=2, column=1, value='Iteration #')

    for col_idx, column in enumerate(columns_data, start=2):
        sheet_transcripts.cell(row=1, column=col_idx, value=column['chat'])
        sheet_transcripts.cell(row=2, column=col_idx, value=column['iteration'])

    # Fill messages and apply formatting
    for col_idx, column in enumerate(columns_data, start=2):
        for msg_idx, message in enumerate(column['messages'], start=3):
            cell = sheet_transcripts.cell(row=msg_idx, column=col_idx, value=message)
            cell.alignment = Alignment(wrap_text=True)

            # Boldface the rating cell
            if column['rating_pos'] and msg_idx == column['rating_pos']:
                cell.font = Font(bold=True)

    # Set column widths and wrap text
    sheet_transcripts.column_dimensions['A'].width = 15  # Width for the first column
    for col in sheet_transcripts.iter_cols(min_col=2, max_col=sheet_transcripts.max_column):
        col_letter = col[0].column_letter
        sheet_transcripts.column_dimensions[col_letter].width = 75
        for cell in col:
            cell.alignment = Alignment(wrap_text=True)

    # Add Plot Image Sheet
    if plot_base64:
        img_sheet_name = "Analysis Plot"
        # Decode the base64 image
        img_data = base64.b64decode(plot_base64)
        image_stream = io.BytesIO(img_data)

        # Create an image object for openpyxl
        img = OpenPyXLImage(image_stream)
        img.width, img.height = img.width * 0.5, img.height * 0.5  # Adjust image size if necessary

        # Add a new sheet for the plot image
        worksheet_plot = workbook.create_sheet(title=img_sheet_name)
        # Insert the image into the sheet
        worksheet_plot.add_image(img, 'A1')

    # Save the workbook to a new BytesIO object
    final_output = io.BytesIO()
    workbook.save(final_output)
    final_output.seek(0)

    return final_output
